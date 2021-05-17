#coding=utf-8
import sys

import openpyxl
from bs4 import BeautifulSoup
import requests
import time
import random
from requests import exceptions
from openpyxl import Workbook
excel_name = "网址.xlsx"
wb = Workbook()
#ws1 = wb.active
#ws1.title='site'
DOWNLOAD_URL='https://top.aizhan.com'

reload(sys)                      # reload 才能调用 setdefaultencoding 方法
sys.setdefaultencoding('utf8')  # 设置 'utf-8'

def get_classurl(url):
    return get_html(url)

def get_html(url):
    ip_list = [
        # '182.84.145.153:3256',
        #'124.94.255.120:9999',
        #'115.233.221.139:3128',
        #'49.88.63.180:8000',
        #'119.3.235.101:3128',
        # '27.191.60.244:3256',
        # '106.45.105.155:3256',
        #'116.17.102.157:3128',
        #'60.191.11.249:3128',
        # '36.56.103.214:9999',
        # '49.87.236.167:9999'
        # '106.45.105.127:3256',
        # '218.88.204.101:3256'
        #'59.55.166.156:3256',
        #'27.43.184.104:9999',
        #'203.132.33.27:3128',
        '47.101.213.111:3128',
        '117.187.167.224:3128',
        '218.16.62.152:3128'

    ]

    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.0'}
    test_ip = random.choice(ip_list)
    print(test_ip)
    try:
        #html = requests.get(url,headers=header,timeout=20).text
        html = requests.get(url,proxies={'https':test_ip} ,headers=header,timeout=20).text
       # html.raise_for_status()
    except Exception as e:
        print(e)
        html = '0'
        pass
    return html

def get_classiflylist(html):
    tmpClassiflyDict = {}
    tmpClassiflyDictList = []
    tmpClassiflyDDict = {}
    findmainclasskey = ''
    findNextClassUrl = ''
    target_classiflyNext = 0
    soup = BeautifulSoup(html,'html.parser')
    classifly_list = soup.find('div', attrs={'class': 'cate'})
    if classifly_list is None:
        return None,'0'  #未找到页面元素，需要重新访问
    try:
        for mainclass_list in classifly_list.find_all('div'):
            classfilter= mainclass_list.find('a').text

            if('子分类' == classfilter):
                for _findmainclassvalue in mainclass_list.find_all('li'):
                    tmpClassiflyDict[_findmainclassvalue.find('a').text] = _findmainclassvalue.find('a').get('href')
                tmpClassiflyDictList.append(tmpClassiflyDict)

            if('其它分类' == classfilter):   #unicode(classfilter.string) navigablestring转str
                for _findmainclasskey in mainclass_list.find_all('li'):

                    if (target_classiflyNext):
                        target_classiflyNext = 0
                        findNextClassUrl = _findmainclasskey.find('a').get('href')  #下一主分类地址
                        break

                    if(_findmainclasskey.has_attr('class')):
                        if('on' in _findmainclasskey.get('class')):
                            target_classiflyNext = 1
                            findmainclasskey = _findmainclasskey.find('a').text
                            tmpClassiflyDDict[findmainclasskey] = tmpClassiflyDictList #主分类字典结构
                if(target_classiflyNext):
                     findNextClassUrl = '1' #未触发下一项结束抓取

    except Exception as e:
        print(e)
        findNextClassUrl = '0'
        pass

    return tmpClassiflyDDict, findNextClassUrl

def get_con(html):
    target_nextpage = 0
    soup = BeautifulSoup(html,'html.parser')
    book_list = soup.find('div', attrs={'class': 'list'})
    if book_list is None:
        return None,None,None,'0'
    #page = soup.find('div', attrs={'class': 'bm bw0 pgs cl'})
    try:
        linkpage_list = soup.find('div', attrs={'class': 'page'})
        for np in linkpage_list.find_all('li'):
            if(target_nextpage):
                target_nextpage = 0
                if(np.has_attr('class')):
                    if("disabled" in np.get('class')):
                        next_page = '1'                   #最后一页结束抓取
                        break
                next_page = np.contents[1].get('href')
                break
            if(np.has_attr('class')):
                if("on" in np.get('class')):
                    target_nextpage = 1
    except Exception as e:
        print(e)
        next_page='0'
        pass

    sitelink = [] #网址
    sitename = []  #网站名
    counttop= [] #总排名
    sitelist = []

    for i in book_list.find_all('li'):
        detail = i.find('div', attrs={'class': 'text'}).find('h2')
        _sitename = detail.find('a').get_text()
        _sitelink = detail.find('em').get_text()
        #sitelink.append(_sitelink)
        #sitename.append(_sitename)

        topcontext = i.find('div', attrs={'class': 'rank'}).find('div', attrs={'class': 'bot clearfix'})
        for _counttopcount in topcontext.find_all('dd'):
            _counttop = unicode(_counttopcount.contents[0])
        #counttop.append(_counttop)
        tmpsitelist = [_sitelink,_sitename,_counttop]
        sitelist.append(tmpsitelist)
    if next_page:
        if '0' == next_page or '1' == next_page:
            return sitelist, next_page
        else:
             return sitelist, DOWNLOAD_URL + next_page
    else:
        return sitelist, None

def get_listvalue(url):
    returnlist = []
    html = '0'
    targetcount = 0
    while url:
        url_old = url
        while '0' == html:
            html = get_html(url)

        tmplist, url = get_con(html)
        targetcount = targetcount + 1
        if 9 == targetcount:
            url = '1'

        if '0' == url:
            url = url_old
        elif '1' == url:
            returnlist.append(tmplist)
            return returnlist
            break  # 结束
        else:
            returnlist.append(tmplist)
            #sitelink = sitelink + _siteLink
            #sitename = sitename + _siteName
            #counttop = counttop + _countTop
        html = '0'
        print(url)



def main():

    url = DOWNLOAD_URL + '/top/t3/'
    classiflyurl = url
    sitelink = [] #网址
    sitename = []  #网站名
    counttop= [] #总排名
    html='0'
    classiflyDDict = {}

    classiflyDir = [
                    #{'休闲娱乐':[{'休闲娱乐':'https://top.aizhan.com/top/t3/'},{'视频电影':'https://top.aizhan.com/top/t3-15/'},{'直播TV':'https://top.aizhan.com/top/t3-429/'}]},
                    #{'生活服务':[{'生活服务':'https://top.aizhan.com/top/t25/'},{'餐饮美食':'https://top.aizhan.com/top/t25-43/'},{'求职招聘':'https://top.aizhan.com/top/t25-45/'}]}
                   ]

#    wb = openpyxl.load_workbook('test.xlsx')

    '''
    showlist = classiflyDir[0]
    tmpkey = showlist.keys()
    tkey = unicode(tmpkey[0])
    showlistkey = showlist.get('休闲娱乐')
    showlistvalue = showlist.values()
    for showlist0 in showlistkey:
        showlist1 = showlist0.get('休闲娱乐')
        showdict = showlist0.get('视频电影')

    testvalue = [["www.t.com","xx","1"],["w.b.com","bb","2"],["ww.c.com","cc","3"]]
    testsvalue = []
    testsvalue.append('test')
    try:
        tlist = []
    #wb = openpyxl.load_workbook('test.xlsx')
        ws = wb.create_sheet(title=unicode('测试'))
        for i in range(1, 5):
            ws = wb.create_sheet(title=str(i))
            #wb.active = 1
            for wvalue in testvalue:
                testlist = ['1', '2', '3']
                tlist.append(testlist)
                ws.append(testsvalue)
        wb.save('test.xlsx')
        wb.close()
    except Exception as e:
        print(e)
        pass
    '''

    while classiflyurl:
        classiflyurl_old = classiflyurl
        while '0' == html:
            html = get_classurl(classiflyurl)

        classiflyDDict, classiflyurl = get_classiflylist(html)
        if '0' == classiflyurl:
            classiflyurl =  classiflyurl_old
        elif '1' == classiflyurl:
            classiflyDir.append(classiflyDDict)
            html = '0'
            break
        else:
            classiflyDir.append(classiflyDDict)
        html = '0'

    try:
        for testlist in classiflyDir:
            testdictkey = testlist.keys()  # 主分类key
            ws = wb.create_sheet(title=unicode(testdictkey[0]))
            for testdict in testlist.get(testdictkey[0]):
                for testdictkeysin in testdict.keys():  # 子分类key
                    tmpdictkeyin = [testdictkeysin]
                    ws.append(tmpdictkeyin)
                    print str(testdictkeysin).decode("utf8")
                    #if(u'男性网站' != testdictkeysin):
                    #    continue
                    testdictinurl = testdict.get(testdictkeysin)  # 子分类url
                    testlistvalue = get_listvalue(testdictinurl)
                    for _testlistvaluecount in testlistvalue:
                        for _testlistvaluetuple in _testlistvaluecount:
                            ws.append(_testlistvaluetuple)
        wb.save('test.xlsx')
        wb.close()
    except Exception as e:
        print(e)
        pass


'''
    while url:
        url_old = url
        while '0' == html:
            html = get_html(url)

        _siteLink,_siteName,_countTop, url = get_con(html)
        if '0' == url:
            url = url_old
        elif '1' == url:
            break            #结束
        else:
            sitelink = sitelink + _siteLink
            sitename = sitename + _siteName
            counttop = counttop + _countTop
        html='0'
        print(url)
        
    for i,j,y in zip(sitelink,sitename,counttop):
        col_A = 'A%s'%(sitelink.index(i)+1)
        col_B = 'B%s'%(sitelink.index(i)+1)
        col_C = 'C%s'%(sitelink.index(i)+1)
        print(i,j,y)
        print(col_A,col_B,col_C)
        #ws1[col_A]=i
        #ws1[col_B]=j
        #ws1[col_C]=y
    wb.save(filename=excel_name)
'''


if __name__ == '__main__':
    main()

